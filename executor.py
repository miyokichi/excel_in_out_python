"""Evaluate an AST node against cell data loaded from an Excel file."""

import re
from typing import Any

import openpyxl

from excel_functions import REGISTRY


# ── Cell address helpers ───────────────────────────────────────────────────────

_COL_RE = re.compile(r"\$?([A-Za-z]+)\$?(\d+)")


def _col_letter_to_index(col: str) -> int:
    col = col.upper().lstrip("$")
    result = 0
    for ch in col:
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


def _parse_address(address: str) -> tuple[int, int]:
    """Return (row, col) as 1-based integers."""
    m = _COL_RE.match(address)
    if not m:
        raise ValueError(f"Invalid cell address: {address!r}")
    col = _col_letter_to_index(m.group(1))
    row = int(m.group(2))
    return row, col


# ── Context: holds all cell values from the workbook ─────────────────────────

class WorkbookContext:
    def __init__(self, excel_path: str, default_sheet: str | None = None) -> None:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        self._sheets: dict[str, Any] = {name: wb[name] for name in wb.sheetnames}
        self.default_sheet = default_sheet or wb.sheetnames[0]

    def get_cell(self, address: str, sheet: str | None = None) -> Any:
        ws = self._sheets[sheet or self.default_sheet]
        return ws[address].value

    def get_range(self, start: str, end: str, sheet: str | None = None) -> list[list[Any]]:
        ws = self._sheets[sheet or self.default_sheet]
        r1, c1 = _parse_address(start)
        r2, c2 = _parse_address(end)
        rows = []
        for r in range(r1, r2 + 1):
            row = []
            for c in range(c1, c2 + 1):
                row.append(ws.cell(row=r, column=c).value)
            rows.append(row)
        return rows


# ── AST evaluator ─────────────────────────────────────────────────────────────

_BINARY_OPS = {
    "+": lambda a, b: a + b,
    "-": lambda a, b: a - b,
    "*": lambda a, b: a * b,
    "/": lambda a, b: a / b,
    "^": lambda a, b: a ** b,
    "&": lambda a, b: str(a) + str(b),
    "=": lambda a, b: a == b,
    "<>": lambda a, b: a != b,
    "<": lambda a, b: a < b,
    ">": lambda a, b: a > b,
    "<=": lambda a, b: a <= b,
    ">=": lambda a, b: a >= b,
}


def evaluate(node: dict, ctx: WorkbookContext) -> Any:
    """Recursively evaluate an AST node and return its value."""
    t = node["type"]

    if t == "Literal":
        return node["value"]

    if t == "CellRef":
        return ctx.get_cell(node["address"], sheet=node.get("sheet"))

    if t == "Range":
        rows = ctx.get_range(node["start"], node["end"], sheet=node.get("sheet"))
        # Return flat list if single row/col, 2D otherwise
        if len(rows) == 1:
            return rows[0]
        if all(len(r) == 1 for r in rows):
            return [r[0] for r in rows]
        return rows

    if t == "BinaryOp":
        op = node["op"]
        left = evaluate(node["left"], ctx)
        right = evaluate(node["right"], ctx)
        fn = _BINARY_OPS.get(op)
        if fn is None:
            raise ValueError(f"Unknown binary operator: {op!r}")
        # Array broadcasting: apply element-wise when either side is a list
        if isinstance(left, list) or isinstance(right, list):
            flat_l = left if isinstance(left, list) else None
            flat_r = right if isinstance(right, list) else None
            length = len(flat_l if flat_l is not None else flat_r)
            l_items = flat_l if flat_l is not None else [left] * length
            r_items = flat_r if flat_r is not None else [right] * length
            return [fn(lv, rv) for lv, rv in zip(l_items, r_items)]
        return fn(left, right)

    if t == "UnaryOp":
        op = node["op"]
        operand = evaluate(node["operand"], ctx)
        if op == "-":
            return -operand
        if op == "+":
            return +operand
        if op == "%":
            return operand / 100
        raise ValueError(f"Unknown unary operator: {op!r}")

    if t == "Function":
        name = node["name"]
        fn = REGISTRY.get(name)
        if fn is None:
            raise NotImplementedError(f"Function {name!r} is not implemented")

        # Special lazy evaluation for IF / IFERROR / IFNA to avoid side-effect errors
        if name == "IF":
            raw_args = node["args"]
            condition = evaluate(raw_args[0], ctx)
            # Array IF: condition is a list → evaluate element-wise
            if isinstance(condition, list):
                true_vals = evaluate(raw_args[1], ctx) if len(raw_args) > 1 else [True] * len(condition)
                false_vals = evaluate(raw_args[2], ctx) if len(raw_args) > 2 else [False] * len(condition)
                if not isinstance(true_vals, list):
                    true_vals = [true_vals] * len(condition)
                if not isinstance(false_vals, list):
                    false_vals = [false_vals] * len(condition)
                return [tv if c else fv for c, tv, fv in zip(condition, true_vals, false_vals)]
            if condition:
                return evaluate(raw_args[1], ctx) if len(raw_args) > 1 else True
            return evaluate(raw_args[2], ctx) if len(raw_args) > 2 else False

        if name in ("IFERROR", "IFNA"):
            try:
                return evaluate(node["args"][0], ctx)
            except Exception:
                return evaluate(node["args"][1], ctx) if len(node["args"]) > 1 else None

        args = [evaluate(arg, ctx) for arg in node["args"]]
        return fn(args)

    if t == "Array":
        return [[evaluate(cell, ctx) for cell in row] for row in node["elements"]]

    if t == "NamedRange":
        raise NotImplementedError(f"Named range {node['name']!r} evaluation not supported without full workbook context")

    raise ValueError(f"Unknown AST node type: {t!r}")


# ── Convenience entry point ───────────────────────────────────────────────────

def run(ast: dict, excel_path: str, sheet: str | None = None) -> Any:
    """Evaluate a single AST against the given Excel file."""
    ctx = WorkbookContext(excel_path, default_sheet=sheet)
    return evaluate(ast, ctx)
