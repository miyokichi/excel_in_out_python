"""Extract formulas and cached values from Excel workbooks."""

import openpyxl
from typing import Any


def extract(path: str) -> dict[str, dict[str, dict[str, Any]]]:
    """Return {sheet_name: {cell_address: {formula, cached_value}}} for all formula cells."""
    wb_formula = openpyxl.load_workbook(path, data_only=False)
    wb_values = openpyxl.load_workbook(path, data_only=True)

    result: dict[str, dict[str, dict[str, Any]]] = {}

    for sheet_name in wb_formula.sheetnames:
        ws_formula = wb_formula[sheet_name]
        ws_values = wb_values[sheet_name]
        cells: dict[str, dict[str, Any]] = {}

        for row in ws_formula.iter_rows():
            for cell in row:
                value = cell.value
                formula_str = None
                if isinstance(value, str) and value.startswith("="):
                    formula_str = value
                elif hasattr(value, "text"):  # ArrayFormula object
                    formula_str = value.text
                if formula_str is not None:
                    address = cell.coordinate
                    cached = ws_values[address].value
                    cells[address] = {"formula": formula_str, "cached_value": cached}

        if cells:
            result[sheet_name] = cells

    return result
